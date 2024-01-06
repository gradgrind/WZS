#deprecated: now concentrating on ods files – see tables.ods_support,
# and grades, ods_template, which uses it.
"""
tables/matrix.py - last updated 2023-12-28

Edit a table template (xlsx).

The base class is <Table>.

<KlassMatrix> handles writing to xlsx templates for pupil-subject matrices.


=+LICENCE=============================
Copyright 2023 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
=-LICENCE========================================
"""

from core.base import Tr
T = Tr("tables.matrix")

### +++++

from typing import Optional

import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter  # , column_index_from_string
from openpyxl.styles import Protection

from core.base import REPORT_ERROR#, REPORT_WARNING

class MatrixError(Exception):
    pass

### -----


class Table:
    """openpyxl based spreadsheet handler ('.xlsx'-files).
    """

    @staticmethod
    def columnLetter(i: int) -> str:
        """Return the column letter(s) for the given (0-based) index."""
        return get_column_letter(i + 1)

    def __init__(self, filepath: str) -> None:
        if not filepath.endswith(".xlsx"):
            filepath += ".xlsx"
        self.template: str = filepath
        self._wb = load_workbook(self.template)
        self.rows: list[list[str]] = []
        for row in self._wb.active.iter_rows():
            values: list[str] = []
            for cell in row:
                v = cell.value
                if isinstance(v, datetime.datetime):
                    v = v.strftime("%Y-%m-%d")
                elif isinstance(v, str):
                    v = v.strip()
                elif v is None:
                    v = ""
                else:
                    v = str(v)
                values.append(v)
            self.rows.append(values)
        self.protected = Protection(locked=True)
        self.unprotected = Protection(locked=False)

    #    def getCell(self, celltag: str) -> Any:
    #        return self._wb.active[celltag].value

    #    def read(self, row: int, col: int) -> Any:
    #        """Read the cell at the given position (0-based indexes).
    #        """
    #        return self.getCell(f"{self.columnLetter(col)}{row+1}")

    def setCell(self, celltag: str, value: str) -> None:
        cell = self._wb.active[celltag]
        cell.value = value
        # To set some "style" (see openpyxl docs):
        # cell.style = style

    def write(
        self,
        row: int,
        col: int,
        val: str,
        protect: Optional[bool] = None
    ) -> None:
        """Write to the cell at the given position (0-based indexes).
        """
        cell = self._wb.active[f"{self.columnLetter(col)}{row+1}"]
        cell.value = val
        # To set/clear cell protection
        if protect is not None:
            cell.protection = self.protected if protect else self.unprotected

    def delEndCols(self, col0: int) -> None:
        """Delete last columns, starting at index <col0> (0-based).
        """
        ndel: int = len(self.rows[0]) - col0
        if ndel > 0:
            self._wb.active.delete_cols(col0 + 1, ndel)

    def delEndRows(self, row0: int) -> None:
        """Delete last rows, starting at index <row0> (0-based).
        """
        ndel = len(self.rows) - row0
        if ndel > 0:
            self._wb.active.delete_rows(row0 + 1, ndel)

    def protectSheet(self, pw: str = None) -> None:
        if pw:
            self._wb.active.protection.set_password(pw)
        else:
            self._wb.active.protection.enable()

    def save(self, filepath: str) -> str:
        if not filepath.endswith(".xlsx"):
            filepath += ".xlsx"
        self._wb.save(filepath)
        return filepath

    def save_bytes(self) -> bytes:
        virtual_workbook = BytesIO()
        self._wb.save(virtual_workbook)
        return virtual_workbook.getvalue()


class ClassMatrix(Table):
    """An extension of the <Table> class to deal with pupil-subject tables.
    """

    def setTitle(self, title: str) -> None:
        """The title cell is at a fixed position, "B1". "A1" is empty.
        """
        self.setCell("B1", title)

    def setInfo(self, info: dict[str, str]) -> None:
        i: int = 0
        for row in self.rows:
            i += 1
            c0: str = row[0]
            if c0:
                if c0[0] == "+":
                    try:
                        v: str = info.pop(c0)
                    except KeyError:
                        REPORT_ERROR(T("INFO_ITEM_MISSING",
                            key = c0, path = self.template
                        ))
                        v = "?"
                    self.setCell(f"C{i}", v)
                else:
                    # The subject key line
                    break
        if info:
            REPORT_ERROR(
                T("INFO_EXCESS_ITEMS",
                    items = ", ".join(info),
                    path = self.template
                ),
            )
        # <row> is the header row
        self.headers: list[str] = row
        # <i> is the row index of the next row (0-based),
        # initially immediately after the headers
        self.header_rowindex: list[int] = [i - 1]
        while True:
            c0 = self.rows[i][0]
            if c0:
                if c0 == "*":
                    self.header_rowindex.append(i)
                else:
                    break
            i += 1
        self.rowindex: int = i - 1
        # column index for header column iteration
        self.hcol: int = 0

    def hideCol(self, index: int, clearheader: bool = False) -> None:
        """Hide the given column (0-indexed). Optionally clear the subject.
        """
        # At some time this wasn't working. It seems to be fixed now ...
        #REPORT_WARNING("WARNING: Column-hiding not working")
        letter: str = self.columnLetter(index)
        self._wb.active.column_dimensions[letter].hidden = True
        if clearheader:
            # Clear any existing "subject"
            for i in self.header_rowindex:
                self.write(1, index, "")

    def hideHeader0(self):
        """Hide the row containing the first header line – for the case
        when the contents are special keys and thus not so relevant for
        the user.
        """
        row = self.header_rowindex[0] + 1
        self._wb.active.row_dimensions[row].hidden= True

    def nextcol(self) -> int:
        """Iterate over header columns with 'X' in template.
        """
        while True:
            self.hcol += 1
            try:
                if self.headers[self.hcol] == "X":
                    return self.hcol
            except:
                raise MatrixError(T("TOO_FEW_COLUMNS", path = self.template))

    def nextrow(self) -> int:
        """Iterate over pupil rows ('X' in first column).
        """
        while True:
            self.rowindex += 1
            try:
                if self.rows[self.rowindex][0] == "X":
                    return self.rowindex
            except:
                raise MatrixError(T("TOO_FEW_ROWS", path = self.template))
